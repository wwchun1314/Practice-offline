import datetime
import json
import re
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox, ttk
from openpyxl import load_workbook

class QuestionBankApp:
    def __init__(self, root):
        self.root = root
        self.root.title("题库生成器 v1.1（支持 Excel）")
        self.root.geometry("850x650")
        self.root.resizable(True, True)

        self.file_path = tk.StringVar()
        self.template_path = tk.StringVar()
        self.save_path = tk.StringVar()

        self.setup_ui()

    def setup_ui(self):
        frame = ttk.Frame(self.root, padding="10")
        frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        # === 标题 ===
        title_label = ttk.Label(frame, text="题库生成器", font=("微软雅黑", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))

        # === 选择题库文件（支持 txt/xlsx）===
        ttk.Label(frame, text="题库文件（.txt 或 .xlsx）：").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(frame, textvariable=self.file_path, width=50, state="readonly").grid(row=1, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))
        ttk.Button(frame, text="浏览...", command=self.select_file).grid(row=1, column=2, padx=5, pady=5)

        # === 选择HTML模板 ===
        ttk.Label(frame, text="HTML模板文件：").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(frame, textvariable=self.template_path, width=50, state="readonly").grid(row=2, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))
        ttk.Button(frame, text="浏览...", command=self.select_template).grid(row=2, column=2, padx=5, pady=5)

        # === 保存路径（自动加日期前缀）===
        ttk.Label(frame, text="保存为：").grid(row=3, column=0, sticky=tk.W, pady=5)
        ttk.Entry(frame, textvariable=self.save_path, width=50, state="readonly").grid(row=3, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))
        ttk.Button(frame, text="选择...", command=self.select_save_path).grid(row=3, column=2, padx=5, pady=5)

        # === 日志输出框 ===
        ttk.Label(frame, text="运行日志：").grid(row=4, column=0, sticky=tk.W, pady=(10, 5))
        self.log_text = scrolledtext.ScrolledText(frame, height=20, wrap=tk.WORD, state="normal", font=("微软雅黑", 10))
        self.log_text.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)

        # 配置行/列权重
        frame.columnconfigure(1, weight=1)
        frame.rowconfigure(5, weight=1)

        # === 按钮区域 ===
        button_frame = ttk.Frame(frame)
        button_frame.grid(row=6, column=0, columnspan=3, pady=10)

        ttk.Button(button_frame, text="开始生成", command=self.generate).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="清空日志", command=self.clear_log).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="退出", command=self.root.quit).pack(side=tk.LEFT, padx=5)

    def log(self, message):
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.log_text.update_idletasks()

    def select_file(self):
        path = filedialog.askopenfilename(
            title="选择题库文件（支持 .txt 或 .xlsx ）",
            filetypes=[
                ("支持的文件", "*.txt *.xlsx"),
                ("文本文件", "*.txt"),
                ("Excel 文件", "*.xlsx")
            ]
        )
        if path:
            self.file_path.set(path)
            self.log(f"✅ 选择题库文件：{path}")

    def select_template(self):
        path = filedialog.askopenfilename(title="选择HTML模板文件", filetypes=[("HTML Files", "*.html")])
        if path:
            self.template_path.set(path)
            self.log(f"✅ 选择模板文件：{path}")

    def select_save_path(self):
        path = filedialog.asksaveasfilename(
            title="保存为",
            defaultextension=".html",
            filetypes=[("HTML Files", "*.html")]
        )
        if path:
            p = Path(path)
            filename = p.name
            stem = p.stem
            suffix = p.suffix
            # 添加日期前缀
            new_filename = f"{datetime.date.today().strftime('%Y%m%d')}_{filename}"
            save_p = p.parent / new_filename
            self.save_path.set(str(save_p))
            self.log(f"✅ 设置保存路径（自动加日期）：{save_p}")

    def generate(self):
        file_path_str = self.file_path.get()
        template_path_str = self.template_path.get()
        save_path_str = self.save_path.get()

        if not file_path_str:
            messagebox.showwarning("警告", "请先选择题库文件！")
            return
        if not template_path_str:
            messagebox.showwarning("警告", "请先选择HTML模板文件！")
            return
        if not save_path_str:
            messagebox.showwarning("警告", "请设置保存路径！")
            return

        try:
            self.log("🔄 开始处理题库文件...")
            file_path = Path(file_path_str)
            template_path = Path(template_path_str)
            save_path = Path(save_path_str)

            # 读取模板
            with template_path.open("r", encoding="utf-8") as f:
                html_content = f.read()

            results = []

            if file_path.suffix.lower() == '.txt':
                results = self.transform_txt(file_path)
                self.log(f"✅ 成功解析 {len(results)} 道 TXT 题目。")
            elif file_path.suffix.lower() == '.xlsx':
                results = self.transform_excel(file_path)
                self.log(f"✅ 成功解析 {len(results)} 道 Excel 题目。")
            elif file_path.suffix.lower() == '.csv':
                results = self.transform_csv(file_path)
                self.log(f"✅ 成功解析 {len(results)} 道 CSV 题目。")
            else:
                raise ValueError("不支持的文件格式！请使用 .txt 、 .xlsx 或 .csv 文件。")

            # 替换模板中的占位符
            webTitle = save_path.stem
            json_str = json.dumps(results, ensure_ascii=False, indent=2)
            html_content = html_content.replace('[{ 替换 }]', json_str)
            html_content = html_content.replace('网页模板', webTitle)

            # 写入文件
            with save_path.open("w", encoding="utf-8") as f:
                f.write(html_content)

            self.log(f"🎉 成功生成文件：{save_path}")
            messagebox.showinfo("成功", f"题库已生成！\n共 {len(results)} 题。\n路径：{save_path}")

        except Exception as e:
            self.log(f"❌ 错误：{e}")
            messagebox.showerror("错误", f"生成失败：{str(e)}")

    def transform_txt(self,path):
        # === TXT 解析：支持多行题干 ===
        questions = []

        with path.open("r", encoding="utf-8") as f:
            content = f.read().replace('．', '.')  # 全角转半角

        # 分割题目
        PATTERN_TITLE = re.compile(r'(?:^|\n)\s*\d+[\.\。\,\、]\s*')
        raw_questions = PATTERN_TITLE.split(content)
        raw_questions = [q.strip() for q in raw_questions if q.strip()]

        for idx, q_text in enumerate(raw_questions):
            lines = [line.strip() for line in q_text.splitlines() if line.strip()]
            if not lines:
                continue

            body_lines = []
            option_lines = []
            in_options = False

            for line in lines:
                if re.match(r'^[A-G][\.\。\,\、]', line):
                    in_options = True
                    option_lines.append(line)
                elif line.startswith("答案：") or line.startswith("答案:"):
                    break
                elif in_options:
                    break  # 防止选项后插入内容
                else:
                    body_lines.append(line)

            title = ' '.join(body_lines)
            options = []
            for line in option_lines:
                match = re.match(r'^[A-G][\.\。\,\、]\s*(.+)', line)
                if match:
                    options.append(match.group(1).strip())

            answer_match = re.search(r'答案[:：]\s*(.*?)(?=\s*解析[:：]|$)', q_text, re.DOTALL)
            answer = answer_match.group(1).strip() if answer_match else ''

            analysis_match = re.search(r'解析[:：]\s*(.*)', q_text, re.DOTALL)
            analysis = analysis_match.group(1).strip() if analysis_match else ''

            questions.append({
                'id': f"q_{idx+1:04d}",
                'title': title,
                'options': options,
                'answer': answer,
                'analysis': analysis
            })

        return questions
    
    def transform_excel(self, path):
        # === Excel 解析 ===
        questions = []

        wb = load_workbook(path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Excel 文件为空")

        headers = [str(h).strip() if h else '' for h in rows[0]]
        data_rows = rows[1:]

        def find_col(keywords):
            for i, h in enumerate(headers):
                if any(kw in h for kw in keywords):
                    return i
            return None

        title_col = find_col(['题目', '题干', 'Question'])
        ans_col = find_col(['答案', 'Answer'])
        analysis_col = find_col(['解析', 'Explanation'])
        type_col = find_col(['题型', 'Type'])  # 新增：题型列

        option_cols = {}
        for letter in "ABCDEFG":
            col_idx = find_col([letter])
            if col_idx is not None:
                option_cols[letter] = col_idx

        if title_col is None or ans_col is None:
            raise ValueError("Excel 缺少必要列：'题目' 和 '答案' 列必须存在")

        # 判断题答案标准化映射
        true_values = {'正确', '对', 'T', 'True', 'true', '是', 'yes', 'Yes'}
        false_values = {'错误', '错', 'F', 'False', 'false', '否', 'no', 'No'}

        for row_idx, row in enumerate(data_rows, start=1):
            if not row or all(cell is None for cell in row):
                continue

            try:
                title = str(row[title_col]).strip() if row[title_col] is not None else ''
                raw_answer = str(row[ans_col]).strip() if row[ans_col] is not None else ''

                # 判断是否为判断题
                is_judgment = False
                if type_col is not None and row[type_col] is not None:
                    type_val = str(row[type_col]).strip()
                    if any(kw in type_val for kw in ['判断', '判断题', 'True/False', 'TF']):
                        is_judgment = True

                # 如果没有题型列，但答案明显是判断类，也可尝试推断（可选）
                # 这里我们保守处理：仅当题型列为判断题时才启用
                options = []
                answer = raw_answer  # 默认保留原答案

                if is_judgment:
                    # 强制设置选项
                    options = ["正确", "错误"]
                    # 映射答案到 A/B
                    if raw_answer in true_values:
                        answer = "A"
                    elif raw_answer in false_values:
                        answer = "B"
                    else:
                        # 如果无法识别，保留原答案，但记录警告（可选）
                        self.log(f"⚠️ 第 {row_idx + 1} 行：无法识别判断题答案 '{raw_answer}'，保留原值")
                else:
                    # 非判断题：按原逻辑读取选项
                    for letter in "ABCDE":
                        if letter in option_cols:
                            val = row[option_cols[letter]]
                            opt = str(val).strip() if val is not None else ''
                            options.append(opt)

                    # 清理末尾空选项
                    while options and not options[-1]:
                        options.pop()

                analysis = ''
                if analysis_col is not None and row[analysis_col] is not None:
                    analysis = str(row[analysis_col]).strip()

                questions.append({
                    'id': f"q_{row_idx:04d}",
                    'title': title,
                    'options': options,
                    'answer': answer,
                    'analysis': analysis
                })

            except Exception as e:
                self.log(f"⚠️ 跳过第 {row_idx + 1} 行（解析错误）: {e}")
                continue

        wb.close()
        return questions

    def clear_log(self):
        self.log_text.delete(1.0, tk.END)

# ============ 启动应用 ============
if __name__ == "__main__":
    root = tk.Tk()
    app = QuestionBankApp(root)
    root.mainloop()
